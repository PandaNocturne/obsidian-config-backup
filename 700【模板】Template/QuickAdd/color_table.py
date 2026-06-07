import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Define the colors from the CSS file (light theme)
colors = {
    '红色': '#ffa8a8',
    '粉色': '#faa2c1',
    '葡萄紫': '#e599f7',
    '紫罗兰': '#b197fc',
    '靛蓝': '#91a7ff',
    '蓝色': '#74c0fc',
    '青色': '#66d9e8',
    '蓝绿': '#63e6be',
    '绿色': '#8ce99a',
    '青柠': '#c0eb75',
    '黄色': '#ffe066',
    '橙色': '#ffc078'
}

# Create a DataFrame with 3x4 layout
df = pd.DataFrame([
    ['红色', '粉色', '葡萄紫', '紫罗兰'],
    ['靛蓝', '蓝色', '青色', '蓝绿'],
    ['绿色', '青柠', '黄色', '橙色']
])

# Create a new Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "多彩表格样式"

# Set column width
for col in range(1, 5):
    ws.column_dimensions[get_column_letter(col)].width = 15

# Set row height
for row in range(1, 4):
    ws.row_dimensions[row].height = 30

# Write data and apply colors
for row_idx, row in enumerate(df.values, 1):
    for col_idx, cell_value in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = cell_value

        # Apply color fill
        color = colors[cell_value]
        # Convert hex to RGB
        rgb = tuple(int(color[i:i+2], 16) for i in (1, 3, 5))
        fill = PatternFill(
            start_color=color[1:], end_color=color[1:], fill_type='solid')
        cell.fill = fill

        # Set font and alignment
        cell.font = Font(size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Save the Excel file
wb.save('color_table.xlsx')
